import openpyxl


TEMPLATE_PATH = 'popular_packages_11/template{0}.xlsx'
wb = openpyxl.load_workbook(TEMPLATE_PATH.format(''))

print(wb.sheetnames)
print(wb['Sheet1'])

active_sheet = wb.active
print(wb.active)

wb.create_sheet('Sheet3')
print(wb.sheetnames)

sheet3 = wb['Sheet3']
sheet3.title = 'NewTitleS3'

print(
    active_sheet.max_row,
    active_sheet.max_column
)

cell = active_sheet['A1']
cell.value = 'FullName'  # update cell value
print(cell.value)

cell2 = active_sheet.cell(row=2, column=1)
print(cell2.value, cell2.row, cell2.column, cell2.coordinate, sep=', ')

print('\n\n')
for row in range(1, active_sheet.max_row + 1):
    full_row = ''
    for col in range(1, active_sheet.max_column + 1):
        cell = active_sheet.cell(row=row, column=col)
        full_row += str(cell.value) + '\t'
    print(full_row)


# Rows
print('\n')
row = active_sheet['1']
print(row)

# Columns
print('\n')
column = active_sheet['A']
print(column)


# Add values
active_sheet.append([1, 2, 3])
print(active_sheet.rows)

# Delete
active_sheet.delete_rows(2, 1)

# Save file
wb.save(TEMPLATE_PATH.format('_mod'))
print('> New file saved')
